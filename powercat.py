from collections.abc import Generator
import sys
import logging
import time
import argparse
from pathlib import Path
from typing import Dict, List, Optional, TypedDict

from langchain_ollama import ChatOllama
from pydantic import BaseModel, Field

from langchain_core.documents import Document
from langchain_core.prompts import ChatPromptTemplate
from langchain_ollama.embeddings import OllamaEmbeddings
from langchain_core.vectorstores import InMemoryVectorStore

from langgraph.graph import START, StateGraph

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from slugify import slugify

from autocat import AutoCat
from consts import Account, Transaction, transaction_cols


logging.basicConfig(level=logging.INFO)
logging.getLogger("httpcore").setLevel(logging.WARN)
logging.getLogger("httpx").setLevel(logging.WARN)
log = logging.getLogger(__name__)


####################################
# CLI

class Args:
    file_path: str
    embedding_model: str
    model: str


parser = argparse.ArgumentParser(usage="Categorize Tiller Money Feeds Transactions")
parser.add_argument(
    "file_path", type=str, help="The path of the excel file to operate on"
)
parser.add_argument(
    "--embedding-model",
    type=str,
    default="nomic-embed-text",
    help="The model name to use for embedding",
)
parser.add_argument(
    "--model", type=str, default="mistral", help="The model name to use for llm"
)
args: Args = parser.parse_args()


####################################
# Const

embedding_model = args.embedding_model
model = args.model

# Minimum reported confidence to accept a categorization
confidence_cuttoff = 80
# High level of confidence, which will skip retries if reported
high_confidence_level = 97
# Maximum number of retries to categorize
max_attempts = 5


llm = ChatOllama(model=model, keep_alive="10m")
embeddings = OllamaEmbeddings(model=embedding_model, keep_alive=10 * 60)
vector_store = InMemoryVectorStore(embeddings)

####################################
# Load workbook


def extract_worksheet_vals(
    sheet: Worksheet, pick_cols: Optional[List[str]] = None
) -> Generator[Dict[str, Cell]]:
    headers: List[str] = []
    col_mapping: Dict[str, Cell] = {}

    for col in sheet.columns:
        header_cell = col[0]
        if pick_cols:
            if header_cell.value in pick_cols:
                headers.append(header_cell.value)
                col_mapping[header_cell.value] = header_cell.column
        elif header_cell.value is not None:
            headers.append(header_cell.value)
            col_mapping[header_cell.value] = header_cell.column

    if pick_cols:
        for col in pick_cols:
            if col not in col_mapping:
                raise Exception(f"Missing column {col} in worksheet {sheet}")

    for idx, row in enumerate(sheet.rows):
        if idx == 0:
            continue

        cell_mapping = {}
        has_value = False
        for column_name, col_idx in col_mapping.items():
            cell = row[col_idx - 1]
            cell_mapping[column_name] = cell
            if cell.value is not None:
                has_value = True

        if has_value:
            yield cell_mapping


def extract_cell_dict(cell_mapping: Dict[str, Cell]) -> Dict[str, any]:
    return {k: c.value for k, c in cell_mapping.items()}


file_path = Path(args.file_path)

if not file_path.exists():
    log.error(f"Provided file {args.file_path} does not exist.")
    sys.exit(1)


workbook = load_workbook(file_path)

accounts_sheet = workbook["Accounts"]
categories_sheet = workbook["Categories"]
transactions_sheet = workbook["Transactions"]

account_cols = ["Account", "Class Override", "Group"]
accounts: List[Account] = [
    extract_cell_dict(d) for d in extract_worksheet_vals(accounts_sheet, account_cols)
]

category_cols = ["Category"]
categories = [
    extract_cell_dict(d)
    for d in extract_worksheet_vals(categories_sheet, category_cols)
]
cats: List[str] = [d["Category"] for d in categories]


transactions = list(extract_worksheet_vals(transactions_sheet, transaction_cols))
transaction_data: List[Transaction] = [
    extract_cell_dict(t)
    for t in extract_worksheet_vals(transactions_sheet, transaction_cols)
]

autocat: Optional[AutoCat] = None

if "AutoCat" in workbook.sheetnames:
    autocat = AutoCat(workbook["AutoCat"])


####################################
# LLM Crap

class Categorization(BaseModel):
    """
    Categorization is the model that is returned from the llm "function call".
    """
    transaction_id: str = Field(
        description="The transaction id from the supplied transaction"
    )
    account_id: str = Field(description="The account id from the supplied transaction")
    description: str = Field(
        description="The description from the supplied transaction"
    )
    category: str = Field(
        description="The category of the transaction, from the supplied list of categories",
        enum=cats,
    )
    confidence_level: int = Field(
        description="The confidence level (in percent between 0 and 100) in your determination of the category",
        gt=0,
        le=100,
    )
    explanation: str = Field(
        description="""A brief explanation as to how you made the categorization determination.
 Also explain how you arrived at your chosen confidence level."""
    )


class State(TypedDict):
    """State is the langchain operation graph state"""
    transaction: Transaction
    previous_transactions: List[Document]
    categorization: Categorization


def transaction_identity(t: Transaction):
    description_slug = slugify(t["Description"])
    return (t["Account"], description_slug)


def get_categorized_transactions(all_transactions: List[Transaction]) -> List[Transaction]:

    examples = []
    seen = set()

    for t in all_transactions:
        if not t.get("Category"):
            continue

        ident = transaction_identity(t)

        if ident in seen:
            continue

        seen.add(ident)

        examples.append(t)

    return examples


def format_transaction_value(col: str, val: any) -> str:
    if col in ("Date", "Date Added", "Month", "Week"):
        fmt = val.date().isoformat()
    else:
        fmt = str(val)

    return f"{col}: '{fmt}'"


def format_transaction(t: Transaction) -> str:
    return "\n".join((format_transaction_value(k, v) for k, v in t.items() if k and v))


def invoke_llm_with_retries(tx: Transaction) -> Optional[Categorization]:
    max_cat: Categorization = None
    for i in range(max_attempts):
        try:
            graph = (
                StateGraph(State)
                .add_sequence([retrieve_previous_txs, categorize])
                .add_edge(START, "retrieve_previous_txs")
                .compile()
            )

            response: State = graph.invoke({"transaction": tx})

            cat = massage_categorization(response["categorization"])

            if cat.confidence_level >= high_confidence_level:
                log.info(f"Categorized with high confidence after {i+1} attempts")
                return cat
            elif cat.confidence_level >= confidence_cuttoff:
                if not max_cat:
                    max_cat = cat
                elif cat.confidence_level > max_cat.confidence_level:
                    max_cat = cat
        except Exception as e:
            log.info(f"Graph failed: {e}")
            continue

    if not max_cat:
        log.error(f"Failed to get data after {max_attempts} attempts.")

    return max_cat


def retrieve_previous_txs(state: State):
    prev_tx_tmpl = ChatPromptTemplate.from_template(
        """Which category was the transaction with the description '{description}',
account '{account}', and institution '{institution}' categorized as?
Say there were no previous categorizations if there were none."""
    )

    similarity_prompt = prev_tx_tmpl.invoke(
        {
            "description": state["transaction"]["Description"],
            "account": state["transaction"]["Account"],
            "institution": state["transaction"]["Institution"],
        }
    )

    prompt = similarity_prompt.to_messages()[0].content

    log.debug(f"Retrieve previous categorizations prompt {prompt}")

    context_res = vector_store.similarity_search(prompt)

    log.debug(f"Retreived {len(context_res)} similar transactions for context")

    return {"previous_transactions": context_res}


def massage_categorization(c: Categorization):
    if c.confidence_level == 1:
        c.confidence_level = 100

    if c.confidence_level < 1 and c.confidence_level > 0:
        c.confidence_level = int(c.confidence_level * 100)

    return c


def categorize(state: State):
    template = ChatPromptTemplate.from_template(
        """Categorize the following transaction.

Here's the list of accounts that the transaction may have come from for context:

```
{accounts}
```

Here is a possibly related previous categorizations, if this categorization matches the
transaction that is provided, you should prefer this previous categorization over 
inferring them from the transaction's fields:

```
{previous_transactions}
```

And here is the transaction that needs to be categorized:

```
{transaction_text}
```
"""
    )

    accts_text = "\n\n".join(
        (
            f"Account Name: '{a["Account"]}', Account Type: '{a["Group"]}', Account Class: '{a["Class Override"]}'"
            for a in accounts
        )
    )

    transaction_text = format_transaction(state["transaction"])
    previous_tx_text = "\n\n".join(
        (", ".join(c.page_content.split("\n")) for c in state["previous_transactions"])
    )

    structured_llm = llm.with_structured_output(Categorization)

    prompt = template.invoke(
        {
            "accounts": accts_text,
            "previous_transactions": previous_tx_text,
            "transaction_text": transaction_text,
        }
    )

    log.debug(f"The full prompt: {prompt.to_messages()[0].content}")

    response = structured_llm.invoke(prompt)

    return {"categorization": response}


start_embedding = time.time()

log.info(f"Begin indexing previous transactions")

categorized_txs = get_categorized_transactions(transaction_data)

example_tx_docs = [
    Document(
        id=t["Transaction ID"],
        page_content=format_transaction(t),
        metadata={"source": "transaction_file"},
    )
    for t in categorized_txs
]

docs = vector_store.add_documents(
    documents=example_tx_docs,
)

# Maps transactions by their (Account, Description)
transaction_map = {
    transaction_identity(t): t["Category"]
    for t in categorized_txs
}

log.info(
    f"Completed indexing transactions in {round(time.time() - start_embedding, 2)}s"
)

start = time.time()
processed = 0
for tx_cells in transactions:

    tx: Transaction = extract_cell_dict(tx_cells)
    if tx["Category"]:
        log.debug("Line is already categorized, skipping.")
        continue

    cat_cell = tx_cells["Category"]

    tx_ident = transaction_identity(tx)
    cat = transaction_map.get(tx_ident)
    if cat:
        log.info(
            f"Categorized based on previous transaction {tx["Description"]} -- {cat}"
        )
        cat_cell.value = cat
        cat_cell.fill = PatternFill(
            start_color="C24AFF", end_color="C24AFF", fill_type="solid"
        )
        continue

    if autocat:
        cat = autocat.try_categorize(tx)
        if cat:
            log.info(
                f"AutoCat categorized {tx["Description"]} -- {cat}"
            )
            cat_cell.value = cat
            cat_cell.fill = PatternFill(
                start_color="4AC8FF", end_color="4AC8FF", fill_type="solid"
            )
            continue

    line_start = time.time()

    log.info("Categorizing line")

    cat = invoke_llm_with_retries(tx)
    category = tx["Category"]

    if cat:
        line_end = time.time()
        log.info(
            f"Categorized ({round(line_end-line_start, 3)} s) {tx["Description"]} -- {cat.category}, {cat.confidence_level} -- {cat.explanation}"
        )
        cat_cell.value = cat.category
        cat_cell.fill = PatternFill(
            start_color="FFF66C", end_color="FFF66C", fill_type="solid"
        )
    else:
        log.info(f"Failed to categorize transaction {tx}")

    processed += 1


log.info(f"Processed {processed} in {time.time() - start} seconds")

workbook.save(file_path)

log.info(f"Workbook saved")
